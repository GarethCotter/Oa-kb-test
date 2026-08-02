"""Audit every editorial internal link in the built articles.

    py scripts/audit-internal-links.py            # summary to stdout
    py scripts/audit-internal-links.py --write    # also write the review workbook

The question being answered: does this link offer the reader a genuine onward
journey - "if you want to do X, this article shows you how" - or is it just a
noun turned blue in the middle of a sentence?

Scope: links inside the article prose only. Navigation chrome is excluded on
purpose - the section rail, breadcrumbs, prev/next, the audience tag's "not you?"
link and the generated on-this-page anchors are all navigation, not editorial
choices, and mixing them in would bury the signal.

Signals recorded per link (none is a verdict on its own):

  framing        the sentence gives a reason to click - "if you", "for guidance",
                 "see X for", "to find out how". This is the strongest single
                 indicator that a link is pulling its weight.
  repeat         the same destination is already linked earlier on this page, so
                 this one is unlikely to add anything.
  bare_title     the anchor text is just the destination's title dropped into a
                 sentence, with no framing around it.
  same_section   points at a neighbouring article in the same section, which the
                 section rail already lists down the side of the page.
  in_step        sits inside a numbered or bulleted step, where sending the reader
                 away mid-procedure is usually unhelpful.
"""

import collections
import os
import re
import sys
from html import unescape

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRAMING = re.compile(
    r'\b(if you|if your|for (more|further|guidance|instructions|information|help|details)'
    r'|see |learn (how|more)|find out|read (more|about)|guidance on|instructions (for|on)'
    r'|to (find|learn|see|do) |more (about|on) |details (of|on)|explained|step[- ]by[- ]step'
    r'|you can (find|read|see))\b', re.I)


def strip_tags(s):
    return re.sub(r'\s+', ' ', unescape(re.sub(r'<[^>]+>', ' ', s))).strip()


# Excel rejects control characters outright, and the corpus carries a scattering of
# private-use glyphs from the original Word documents that render as boxes. Strip
# both before anything reaches a cell.
ILLEGAL = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f-]')


def cell_safe(v):
    return ILLEGAL.sub('', v) if isinstance(v, str) else v


def pages():
    for sec in sorted(os.listdir(ROOT)):
        d = os.path.join(ROOT, sec)
        if re.match(r'^\d\d-', sec) and os.path.isdir(d):
            for f in sorted(os.listdir(d)):
                if f.endswith('.html') and f != 'index.html':
                    yield sec, f[:-5], os.path.join(d, f)


def title_of(path):
    if not os.path.exists(path):
        return ''
    m = re.search(r'<h1[^>]*>(.*?)</h1>', open(path, encoding='utf-8').read(), re.S)
    return strip_tags(m.group(1)) if m else ''


def collect():
    titles = {}
    for sec, slug, p in pages():
        titles[(sec, slug)] = title_of(p)

    rows = []
    for sec, slug, p in pages():
        html = open(p, encoding='utf-8').read()
        m = re.search(r'<div class="prose">(.*?)</div>\s*(?:<div class="feedback|<nav class="prevnext)',
                      html, re.S)
        if not m:
            continue
        prose = m.group(1)
        # Drop the generated heading anchors before anything else - they render as a
        # bare "#" and would otherwise pepper every context string.
        prose = re.sub(r'<a\b[^>]*class="anchor"[^>]*>.*?</a>', '', prose, flags=re.S)

        # Build a plain-text version of the page with each editorial link replaced by a
        # marker, so the context can be sliced from TEXT rather than from raw HTML.
        # Slicing HTML by character offset lands mid-tag and leaks attribute fragments
        # like '-set-up-the-event">' into the output, which makes the sheet unscannable.
        marked, targets = prose, []
        for a in re.finditer(r'<a\b([^>]*)href="([^"]+)"([^>]*)>(.*?)</a>', prose, re.S):
            attrs = a.group(1) + a.group(3)
            if 'aud-alt' in attrs or a.group(2).startswith(('http', 'mailto:', '#')):
                continue
            targets.append(a)
        for i, a in enumerate(reversed(targets)):
            n = len(targets) - 1 - i
            marked = marked[:a.start()] + ('\x00%d\x00' % n) + marked[a.end():]
        plain = strip_tags(marked)

        seen = set()
        for n, a in enumerate(targets):
            attrs = a.group(1) + a.group(3)
            href, inner = a.group(2), a.group(4)
            text = strip_tags(inner)
            if not text:
                continue
            tgt = href.split('#')[0].replace('.html', '')
            if tgt.startswith('../'):
                tsec, tslug = tgt[3:].split('/', 1) if '/' in tgt[3:] else (sec, tgt[3:])
            elif tgt.startswith('/'):
                parts = tgt.lstrip('/').split('/')
                tsec, tslug = (parts[0], parts[1]) if len(parts) > 1 else (sec, parts[0])
            else:
                tsec, tslug = sec, tgt
            dest_title = titles.get((tsec, tslug), '')
            if not dest_title and tslug in ('index', ''):
                # a link to a section landing page, not an article
                dest_title = 'SECTION PAGE - ' + tsec

            # sentence around the link, sliced from plain text via the marker
            tok = '\x00%d\x00' % n
            i = plain.find(tok)
            if i < 0:
                before = after = ''
            else:
                before = plain[max(0, i - 200):i]
                after = plain[i + len(tok):i + len(tok) + 140]
            before = re.sub(r'\x00\d+\x00', '…', before)
            after = re.sub(r'\x00\d+\x00', '…', after)
            sent = (before[-150:].lstrip() + ' [' + text + '] ' + after[:90]).strip()

            key = (tsec, tslug)
            rows.append({
                'section': sec,
                'article': titles.get((sec, slug), slug),
                'slug': slug,
                'anchor': text,
                'dest_section': tsec,
                'dest': dest_title or tslug,
                'dest_slug': tslug,
                'context': sent,
                'framing': bool(FRAMING.search(before[-120:] + ' ' + after[:60])),
                'repeat': key in seen,
                'bare_title': bool(dest_title) and text.strip().lower().rstrip('.') == dest_title.strip().lower(),
                'same_section': tsec == sec,
                'in_step': bool(re.search(r'<li[^>]*>(?:(?!</li>).)*$',
                                          prose[max(0, a.start() - 400):a.start()], re.S)),
                'self': (tsec, tslug) == (sec, slug),
            })
            seen.add(key)
    return rows


def verdict(r):
    if r['self']:
        return 'Remove - links to itself'
    if r['repeat']:
        return 'Likely redundant - already linked above'
    if r['framing']:
        return 'Keep - offers a reason to click'
    if r['bare_title'] and r['same_section']:
        return 'Review - bare title, same section (rail already lists it)'
    if r['bare_title']:
        return 'Review - bare title, no framing'
    return 'Review'


def main():
    rows = collect()
    for r in rows:
        r['verdict'] = verdict(r)
    print('editorial internal links in article prose: %d across %d articles\n'
          % (len(rows), len({r['slug'] for r in rows})))
    for k, v in collections.Counter(r['verdict'] for r in rows).most_common():
        print('   %-58s %d' % (k, v))
    print('\nmost-linked destinations:')
    for (d, n) in collections.Counter(r['dest'] for r in rows).most_common(10):
        print('   %-56s %d' % (d[:55], n))

    if '--write' in sys.argv:
        write_workbook(rows)
    return 0


def write_workbook(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    A = 'Arial'
    HF = PatternFill('solid', fgColor='101C38')
    HFONT = Font(name=A, bold=True, color='FFFFFF', size=10)
    BODY = Font(name=A, size=10)
    BOLD = Font(name=A, size=10, bold=True)
    TITLE = Font(name=A, size=14, bold=True, color='101C38')
    MUTED = Font(name=A, size=9, color='595959', italic=True)
    IN = PatternFill('solid', fgColor='FFFF00')
    TH = Side(style='thin', color='D9D9D9')
    BOX = Border(left=TH, right=TH, top=TH, bottom=TH)

    wb = Workbook()

    ws = wb.active
    ws.title = 'How to use'
    ws.sheet_view.showGridLines = False
    intro = [
        ('Internal link review', TITLE), ('', None),
        ('Every editorial link inside the article text of the help centre, with the sentence', BODY),
        ('it sits in. Generated %s.' % __import__('datetime').date.today().isoformat(), BODY),
        ('', None),
        ('The question to ask of each one', BOLD),
        ('Does this link give the reader a reason to click - "if you want to do X, this article', BODY),
        ('shows you how"? Or is it just a noun turned blue in the middle of a sentence?', BODY),
        ('', None),
        ('What is NOT in here', BOLD),
        ('Navigation is excluded on purpose: the section rail, breadcrumbs, prev/next links,', BODY),
        ('the audience tag and the on-this-page box. Those are generated, not editorial', BODY),
        ('choices, and including them would bury the signal.', BODY),
        ('', None),
        ('The columns that do the work', BOLD),
        ('Context      the sentence, with the linked words in [square brackets]', BODY),
        ('Framing      TRUE if the sentence gives a reason to click. Strongest single signal', BODY),
        ('             that the link is earning its place.', BODY),
        ('Repeat       the same destination is already linked higher up the same page', BODY),
        ('Bare title   the anchor is just the destination article title, with no framing', BODY),
        ('Same section the target sits in the same section, which the rail already lists', BODY),
        ('In step      the link is inside a numbered or bulleted step, where sending the', BODY),
        ('             reader away mid-procedure usually hurts', BODY),
        ('', None),
        ('Suggested is a starting point, not a decision. Sort or filter by it, then use', MUTED),
        ('the Decision column. Nothing here has been changed in the site.', MUTED),
        ('', None),
        ('Fill in these two columns (shaded yellow)', BOLD),
    ]
    for i, (t, f) in enumerate(intro, start=1):
        c = ws.cell(row=i, column=1, value=cell_safe(t))
        if f:
            c.font = f
    r = len(intro) + 1
    for j, h in enumerate(['Decision', 'Notes']):
        c = ws.cell(row=r, column=1 + j, value=h)
        c.font = HFONT; c.fill = HF; c.border = BOX
    for j, v in enumerate(['Remove', 'Reader is mid-procedure here; sending them away loses their place']):
        c = ws.cell(row=r + 1, column=1 + j, value=v)
        c.font = BODY; c.fill = IN; c.border = BOX
    ws.cell(row=r + 2, column=1,
            value='Example row. Decision must be Keep, Remove, or Reword.').font = MUTED
    r += 4
    ws.cell(row=r, column=1, value='Progress').font = BOLD
    n = len(rows)
    prog = [('Links to review', '=COUNTA(Links!D2:D%d)' % (n + 1)),
            ('Decision recorded', '=COUNTA(Links!L2:L%d)' % (n + 1)),
            ('Still to do', '=B%d-B%d' % (r + 1, r + 2)),
            ('', None),
            ('Marked Keep', '=COUNTIF(Links!L2:L%d,"Keep")' % (n + 1)),
            ('Marked Remove', '=COUNTIF(Links!L2:L%d,"Remove")' % (n + 1)),
            ('Marked Reword', '=COUNTIF(Links!L2:L%d,"Reword")' % (n + 1))]
    for i, (lab, f) in enumerate(prog, start=1):
        ws.cell(row=r + i, column=1, value=lab).font = BODY
        if f:
            ws.cell(row=r + i, column=2, value=f).font = BOLD
    ws.column_dimensions['A'].width = 82
    ws.column_dimensions['B'].width = 62

    ws = wb.create_sheet('Links')
    head = ['Section', 'Article', 'Anchor text', 'Context (link in [brackets])', 'Links to',
            'Framing', 'Repeat', 'Bare title', 'Same section', 'In step', 'Suggested',
            'Decision', 'Notes']
    for j, h in enumerate(head, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HFONT; c.fill = HF; c.border = BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    rows.sort(key=lambda r: (r['verdict'].startswith('Keep'), r['section'], r['article']))
    for i, r_ in enumerate(rows, start=2):
        vals = [r_['section'], r_['article'], r_['anchor'], r_['context'], r_['dest'],
                r_['framing'], r_['repeat'], r_['bare_title'], r_['same_section'],
                r_['in_step'], r_['verdict']]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=cell_safe(v))
            c.font = BODY; c.border = BOX
            if j in (6, 7, 8, 9, 10):
                c.alignment = Alignment(horizontal='center')
            if j == 4:
                c.alignment = Alignment(wrap_text=True, vertical='top')
        for j in (12, 13):
            c = ws.cell(row=i, column=j); c.fill = IN; c.border = BOX; c.font = BODY
    dv = DataValidation(type='list', formula1='"Keep,Remove,Reword"', allow_blank=True)
    ws.add_data_validation(dv); dv.add('L2:L%d' % (len(rows) + 1))
    for j, w in enumerate([26, 40, 34, 76, 38, 9, 9, 10, 12, 9, 44, 14, 40], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:M%d' % (len(rows) + 1)
    ws.row_dimensions[1].height = 32

    ws = wb.create_sheet('Most linked')
    for j, h in enumerate(['Destination article', 'Times linked', 'Linked from sections'], start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HFONT; c.fill = HF; c.border = BOX
    agg = collections.Counter(r['dest'] for r in rows)
    secs = collections.defaultdict(set)
    for r_ in rows:
        secs[r_['dest']].add(r_['section'])
    for i, (d, n_) in enumerate(agg.most_common(), start=2):
        for j, v in enumerate([d, n_, ', '.join(sorted(secs[d]))], start=1):
            c = ws.cell(row=i, column=j, value=cell_safe(v)); c.font = BODY; c.border = BOX
    for j, w in enumerate([56, 14, 60], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:C%d' % (len(agg) + 1)

    out = os.path.join(ROOT, 'project', 'internal-link-review.xlsx')
    wb.save(out)
    print('\nwrote %s' % os.path.relpath(out, ROOT))


if __name__ == '__main__':
    sys.exit(main() or 0)
